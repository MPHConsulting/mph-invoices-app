"""
Migrate the MPH Consulting invoice workbook into the PWA seed dataset.

Reads three sheets from ``MPH Invoices - PWA.xlsx``:
  * Customers        - the client directory
  * Invoices - Main  - one row per invoice (header + totals)
  * Invoice Details  - one row per line item, joined to an invoice by Invoice #

and writes ``src/data/invoices-data.json`` with 100% numeric fidelity: every
line item's amount is copied verbatim from the sheet's "Total" column, and the
per-invoice subtotal is reconciled against the workbook's own "Detail Total".

Run with:  python migration/migrate.py
"""

from __future__ import annotations

import json
import os
import re
from datetime import datetime, date

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
PROJECT = os.path.dirname(HERE)
# The workbook lives next to the project folder, under WebApps.
WEBAPPS = os.path.dirname(PROJECT)
SOURCE = os.path.join(WEBAPPS, "MPH Consulting Invoices", "MPH Invoices - PWA.xlsx")
# Private full dataset (gitignored). It is NOT bundled into the public app; it is
# only used to seed the user's private cloud-sync Gist via scripts/build_gist_body.py.
OUT = os.path.join(PROJECT, "migration", "invoices-full.json")


def s(v) -> str:
    """Clean cell -> trimmed string ('' for blanks)."""
    if v is None:
        return ""
    if isinstance(v, str):
        return v.strip()
    return str(v)


def num(v) -> float:
    """Cell -> float (0 for blanks / non-numeric)."""
    if v is None or v == "":
        return 0.0
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def iso(v) -> str | None:
    """Cell -> ISO date string, or None."""
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, str) and v.strip():
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%m/%d/%Y"):
            try:
                return datetime.strptime(v.strip(), fmt).strftime("%Y-%m-%d")
            except ValueError:
                pass
    return None


def company_no(label: str) -> int | None:
    """Leading integer of a '2 - Company Name' style label."""
    m = re.match(r"\s*(\d+)\s*-", label)
    return int(m.group(1)) if m else None


def company_name_part(label: str) -> str:
    """Name portion of a '2 - Company Name' style label."""
    return re.sub(r"^\s*\d+\s*-\s*", "", label).strip()


def norm(name: str) -> str:
    """Normalise a company name for fuzzy matching."""
    return re.sub(r"[^a-z0-9]+", " ", name.lower()).strip()


def match_customer(label: str, by_no: dict[int, dict], by_name: dict[str, dict]):
    """Resolve an invoice's company label to a customer id.

    The source workbook has a few data-entry slips where the leading number is
    wrong (e.g. '4 - Offshore Services Australasia' should be customer #2), so we
    prefer an exact name match and only fall back to the number.
    """
    name = norm(company_name_part(label))
    if name and name in by_name:
        return by_name[name]["id"]
    no = company_no(label)
    if no in by_no:
        return by_no[no]["id"]
    return None


def load_customers(wb) -> list[dict]:
    ws = wb["Customers"]
    customers = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        # Columns: _, Lookup, Company #, Name, Contact, Address, City, State, ZIP, Phone, Email, Fax
        company = s(row[3])
        no = row[2]
        if not company or no in (None, ""):
            continue
        try:
            no = int(no)
        except (TypeError, ValueError):
            continue
        customers.append(
            {
                "id": f"cust-{no}",
                "companyNo": no,
                "companyName": company,
                "contactName": s(row[4]),
                "address": s(row[5]),
                "city": s(row[6]),
                "state": s(row[7]),
                "zip": s(row[8]),
                "phone": s(row[9]),
                "email": s(row[10]),
                "fax": s(row[11]),
                "origin": "excel",
            }
        )
    return customers


def load_details(wb) -> dict[str, list[dict]]:
    ws = wb["Invoice Details"]
    by_invoice: dict[str, list[dict]] = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        # Columns: _, Description, Invoice #, Item #, Qty, Unit Price, Discount, Total
        invoice_no = s(row[2])
        desc = s(row[1])
        if not invoice_no or not desc:
            continue
        by_invoice.setdefault(invoice_no, []).append(
            {
                "description": desc,
                "itemNo": s(row[3]),
                "qty": num(row[4]),
                "unitPrice": num(row[5]),
                "discount": num(row[6]),
                # Amount is copied verbatim from the sheet for perfect fidelity.
                "amount": round(num(row[7]), 2),
            }
        )
    return by_invoice


def load_invoices(wb, details: dict[str, list[dict]], customers: list[dict]):
    ws = wb["Invoices - Main"]
    by_no = {c["companyNo"]: c for c in customers}
    by_name = {norm(c["companyName"]): c for c in customers}
    invoices = []
    recon = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        # Columns: _, Invoice #, Company, Date, Project, Tax Rate, Other, Deposit,
        #          Detail Total, Invoice Total, Notes
        invoice_no = s(row[1])
        # Skip blank rows, the repeated header, and the trailing totals row.
        if not invoice_no or invoice_no.lower() in ("totals", "total", "invoice #"):
            continue
        company_label = s(row[2])
        customer_id = match_customer(company_label, by_no, by_name)
        items = details.get(invoice_no, [])
        subtotal = round(sum(i["amount"] for i in items), 2)
        detail_total = round(num(row[8]), 2)
        invoice_total = round(num(row[9]), 2)
        if abs(subtotal - detail_total) > 0.5:
            recon.append(
                {"invoice": invoice_no, "sumOfLines": subtotal, "sheetDetailTotal": detail_total}
            )
        invoices.append(
            {
                "id": f"inv-{invoice_no}",
                "invoiceNo": invoice_no,
                "customerId": customer_id,
                "companyLabel": company_label,
                "date": iso(row[3]),
                "projectDescription": s(row[4]),
                "taxRate": num(row[5]),
                "other": num(row[6]),
                "deposit": num(row[7]),
                "notes": s(row[10]),
                "lineItems": items,
                "importedDetailTotal": detail_total,
                "importedInvoiceTotal": invoice_total,
                "origin": "excel",
            }
        )
    return invoices, recon


def main() -> None:
    if not os.path.exists(SOURCE):
        raise SystemExit(f"Workbook not found: {SOURCE}")
    wb = openpyxl.load_workbook(SOURCE, data_only=True)

    customers = load_customers(wb)
    details = load_details(wb)
    invoices, recon = load_invoices(wb, details, customers)

    # Sort invoices by date (undated last), then by invoice number.
    def sort_key(inv):
        d = inv["date"] or "9999-99-99"
        return (d, inv["invoiceNo"])

    invoices.sort(key=sort_key)

    company_profile = {
        "name": "MPH Consulting",
        "addressLines": "160 Robinson Rd, #14-04\nSingapore Business Federation Centre\nSingapore 068914",
        "phone": "",
        "email": "",
        "bankDetails": "",
        "paymentTerms": "Total due in 30 days. Overdue accounts subject to a service charge of 5% per month.",
        "currency": "AUD",
    }

    data = {
        "meta": {
            "generated": datetime.now().isoformat(timespec="seconds"),
            "source": os.path.basename(SOURCE),
            "numInvoices": len(invoices),
            "numCustomers": len(customers),
            "numLineItems": sum(len(i["lineItems"]) for i in invoices),
        },
        "companyProfile": company_profile,
        "customers": customers,
        "invoices": invoices,
    }

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)

    print(f"Wrote {OUT}")
    print(
        f"  {len(customers)} customers, {len(invoices)} invoices, "
        f"{data['meta']['numLineItems']} line items"
    )
    invoices_without_customer = [i["invoiceNo"] for i in invoices if not i["customerId"]]
    if invoices_without_customer:
        print(f"  invoices with no matched customer: {invoices_without_customer}")
    if recon:
        print(f"  {len(recon)} invoices where line sum != sheet detail total:")
        for r in recon:
            print(f"    {r['invoice']}: lines={r['sumOfLines']} sheet={r['sheetDetailTotal']}")
    else:
        print("  all invoice line-sums reconcile with the sheet detail totals.")


if __name__ == "__main__":
    main()
